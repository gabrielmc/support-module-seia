# app/routes/relatorios_routes.py

import pandas as pd
from io import BytesIO
from shapely import wkb
from openpyxl.styles import Alignment
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from fastapi import UploadFile, File, HTTPException
from openpyxl.styles import Alignment, Font
from app.services.relatorio_service import RelatorioService

router = APIRouter(
    prefix="/relatorios",
    tags=["RELATÓRIO"]
)


@router.get("/zoneamento-UC/{periodo}")
def listar_relatorios(periodo: str):
    try:
        data_fomatada = pd.to_datetime(periodo, dayfirst=True).strftime("%Y-%m-%d")
        dados_relatorios = RelatorioService().relatorio_uc(data_fomatada)
        if not dados_relatorios:
            raise HTTPException(status_code=404, detail="Nenhum dado encontrado")
        
        # Converte para DataFrame
        df = pd.DataFrame(
            dados_relatorios["dados"],
            columns=dados_relatorios["colunas"]
        )     
        #Mapeamento de colunas (interno → solicitado)
        colunas_map = {
            "processo": "Número do processo",
            "requerente": "Nome do requerente",
            "cpf_cnpj": "CPF/CNPJ",
            "nome_empreendimento": "Nome do empreendimento",
            "tipologia_empreendimento": "Tipologia do Empreendimento",
            "municipio_empreendimento": "Município do Empreendimento",
            "ato_processo": "Ato do processo",
            "data_conclusao": "Data de conclusão do processo",
            "status_ato": "Status do ato",
            "validade_ato": "Validade do ato",
            "coordenadas": "Coordenadas geográficas"
        }
        df.rename(columns=colunas_map, inplace=True)
        
        # Reordena as colunas exatamente na ordem solicitada
        ordem_colunas = list(colunas_map.values())
        df = df[[col for col in ordem_colunas if col in df.columns]]
        
        #Cria o arquivo Excel em memória
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Zoneamento UC")

        ws = writer.sheets["Zoneamento UC"]

        # Centraliza todas as células que possuem valor
        for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column
        ):
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True
                    )

        # Ajusta largura das colunas
        for col_cells in ws.columns:
            max_length = max(
                (len(str(cell.value)) for cell in col_cells if cell.value),
                default=0
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_length + 3, 45)

            output.seek(0)
            filename = f"Relatorio_zoneamento_uc_{periodo}.xlsx"
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename={filename}"
                }
            )
    except HTTPException:
        raise
    except Exception as e:
        print(f"Erro ao gerar relatório: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/zoneamento-UC/csv-to-xlsx")
def converter_csv_para_excel(arquivo: UploadFile = File(...)):
    try:
        # Validação básica
        if not arquivo.filename.endswith(".csv"):
            raise HTTPException(status_code=400, detail="Envie um arquivo .csv")

        # Lê o CSV
        df = pd.read_csv(
            arquivo.file,
            sep=";",      # ajuste se necessário
            encoding="utf-8"
        )

        # Mapeamento de colunas (CSV → relatório final)
        colunas_map = {
            "processo": "Número do processo",
            "requerente": "Nome do requerente",
            "cpf_cnpj": "CPF/CNPJ",
            "nome_empreendimento": "Nome do empreendimento",
            "tipologia_empreendimento": "Tipologia do Empreendimento",
            "municipio_empreendimento": "Município do Empreendimento",
            "ato_processo": "Ato do processo",
            "data_conclusao": "Data de conclusão do processo",
            "status_ato": "Status do ato",
            "validade_ato": "Validade do ato",
            "coordenadas": "Coordenadas geográficas"
        }

        # Renomeia colunas
        df.rename(columns=colunas_map, inplace=True)

        # Reordena colunas
        ordem_colunas = list(colunas_map.values())
        df = df[[c for c in ordem_colunas if c in df.columns]]

        # Cria Excel em memória
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Zoneamento UC")
            ws = writer.sheets["Zoneamento UC"]

            # Cabeçalho em negrito
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Centraliza TODO o conteúdo
            for row in ws.iter_rows(
                min_row=1,
                max_row=ws.max_row,
                min_col=1,
                max_col=ws.max_column
            ):
                for cell in row:
                    if cell.value is not None:
                        cell.alignment = Alignment(
                            horizontal="center",
                            vertical="center",
                            wrap_text=True
                        )

            # Ajusta largura das colunas
            for col_cells in ws.columns:
                max_length = max(
                    (len(str(cell.value)) for cell in col_cells if cell.value),
                    default=0
                )
                ws.column_dimensions[
                    col_cells[0].column_letter
                ].width = min(max_length + 3, 45)

        output.seek(0)
        filename = arquivo.filename.replace(".csv", ".xlsx")
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    except Exception as e:
        print(f"Erro ao converter CSV: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/zoneamento-UC/xlsx-corrigir-coordenadas")
def corrigir_xlsx_coordenadas(arquivo: UploadFile = File(...)):
    try:
        if not arquivo.filename.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Envie um arquivo .xlsx")

        # CORREÇÃO CRÍTICA AQUI
        conteudo = arquivo.file.read()
        excel_bytes = BytesIO(conteudo)
        df = pd.read_excel(excel_bytes)
        coluna_coord = "Coordenadas geográficas"
        if coluna_coord not in df.columns:
            raise HTTPException(
                status_code=400,
                detail=f"Coluna '{coluna_coord}' não encontrada no arquivo"
            )
        def wkb_hex_to_wkt(valor):
            if not valor or not isinstance(valor, str):
                return None
            try:
                geom = wkb.loads(bytes.fromhex(valor))
                return geom.wkt
            except Exception:
                return valor

        # Converte WKB HEX → POINT (lon lat)
        df[coluna_coord] = df[coluna_coord].apply(wkb_hex_to_wkt)

        # Cria novo Excel em memória
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Zoneamento UC")
            ws = writer.sheets["Zoneamento UC"]

            # Cabeçalho em negrito
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Centraliza conteúdo
            for row in ws.iter_rows(
                min_row=1,
                max_row=ws.max_row,
                min_col=1,
                max_col=ws.max_column
            ):
                for cell in row:
                    if cell.value is not None:
                        cell.alignment = Alignment(
                            horizontal="center",
                            vertical="center",
                            wrap_text=True
                        )

            # Ajuste de largura
            for col_cells in ws.columns:
                max_length = max(
                    (len(str(cell.value)) for cell in col_cells if cell.value),
                    default=0
                )
                ws.column_dimensions[
                    col_cells[0].column_letter
                ].width = min(max_length + 3, 50)

        output.seek(0)
        filename = arquivo.filename.replace(".xlsx", "_corrigido.xlsx")
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"Erro ao corrigir XLSX: {e}")
        raise HTTPException(status_code=500, detail=str(e))

